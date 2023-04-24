from openpyxl import Workbook
import function

class hr:
    def __init__(self,name,mail,mobile,adress,motive):
        self.name = name
        self.mail = mail
        self.mobile = mobile
        self.adress = adress
        self.motive = motive

input_list = []

for i in range(5):
    input_list.append(str(input()))

print(input_list)

"""
name:Mark
mail:asdfghjk12345@spring-j510.com
mobile:+34-090123412341234
adress:France paris
reason for (one's) desire:As an AI language model
"""

myinstance = hr(input_list[0],input_list[1],input_list[2],input_list[3],input_list[4])
print(myinstance.name)

path = "xlsx/" + function.excel_title()

# ワークブックの新規作成と保存
wb = Workbook()
wb.save(path)

# ワークブックの読み込み
from openpyxl import load_workbook

wb = load_workbook(path)

# 変更したいシート名を指定する
old_sheet_name = 'Sheet'
new_sheet_name = 'summary'

# シート名を変更する
ws = wb[old_sheet_name]
ws.title = new_sheet_name


# ワークシートの選択
ws = wb['summary']  # ワークシートを指定
ws = wb.active  # アクティブなワークシートを選択

# セルに書き込み
ws['B2'] = 'No.'
ws['C2'] = 'name'
ws['D2'] = 'Mail'
ws['E2'] = 'Mobile'
ws['F2'] = 'Adress'
ws['G2'] = 'reason for (ones) desire'

wb.save(path)  # overwrite myworkbook.xlsx